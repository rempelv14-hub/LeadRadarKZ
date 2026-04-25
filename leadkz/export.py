from __future__ import annotations

import csv
import json
import shutil
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from .database import Database
from .filters import detect_city, estimate_budget, reply_risk
from .formatting import dt_to_local_text

COLUMNS = [
    "Дата", "Оценка", "Город", "Категория", "Бюджет", "Риск", "Статус",
    "Группа", "Автор", "Текст", "Причина", "Ссылка", "Дубликаты",
]


def _rows(db: Database, since: datetime):
    for row in db.get_export_rows_since(since):
        text = row["text"]
        category = row["category"]
        yield [
            dt_to_local_text(row["message_date"]), row["score"], detect_city(f"{text} {row['chat_title']}"),
            category, estimate_budget(text, category), reply_risk(text, int(row["score"]), category),
            row["status"], row["chat_title"], row["sender_username"] or "", text,
            row["reasons"], row["link"] or "", row["duplicate_count"],
        ]


def export_leads_csv(db: Database, since: datetime, export_dir: Path) -> Path:
    export_dir.mkdir(parents=True, exist_ok=True)
    path = export_dir / f"leadkz_leads_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(COLUMNS)
        writer.writerows(_rows(db, since))
    return path


def export_leads_xlsx(db: Database, since: datetime, export_dir: Path) -> Path:
    export_dir.mkdir(parents=True, exist_ok=True)
    path = export_dir / f"leadkz_leads_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.append(COLUMNS)
    for row in _rows(db, since):
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
    widths = [18, 10, 20, 28, 36, 42, 16, 28, 20, 70, 55, 35, 12]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(path)
    return path


def export_settings_json(db: Database, export_dir: Path) -> Path:
    export_dir.mkdir(parents=True, exist_ok=True)
    path = export_dir / f"leadkz_settings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    payload = {"exported_at": datetime.now(timezone.utc).isoformat(), "settings": db.get_all_settings()}
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def backup_database(db: Database, export_dir: Path) -> Path:
    export_dir.mkdir(parents=True, exist_ok=True)
    db.conn.commit()
    path = export_dir / f"leadkz_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.sqlite3"
    shutil.copy2(db.path, path)
    return path
